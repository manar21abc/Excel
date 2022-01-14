from openpyxl.workbook import Workbook
from openpyxl import load_workbook

wb = load_workbook("appointments.xlsx")
ws = wb.active
# print(ws.title)

# j represent the raw number of the excel sheet
j = 1

with open("special_format.txt", "w") as f:
    for n in range(1, 27):
        passport_no = ws[f"G{n}"].value
        name = ws[f"J{n}"].value
        surname = ws[f"K{n}"].value

        # I had to add explicitly add 0 because excel ignores the leading 0
        if ws[f"L{n}"].value == 1:
            gender = "0" + str(ws[f"L{n}"].value)
        else:
            gender = ws[f"L{n}"].value

        # slicing is used to extract only the date value (by default it brings the time also)
        expiry = str(ws[f"M{n}"].value)[0:10]
        birthday = str(ws[f"N{n}"].value)[0:10]

        f.write(f'r{j},2,"^document\[00000000{passport_no}\]$","""10""","",0,c119')
        f.write("\n")
        f.write(f'r{j+1},0,"^name\[00000000{passport_no}\]$",{name},"",1,c119')
        f.write("\n")
        f.write(f'r{j+2},0,"^surname\[00000000{passport_no}\]$",{surname},"",1,c119')
        f.write("\n")
        f.write(f'r{j+3},3,"^gender\[00000000{passport_no}\]$",{gender},"",0,c119')
        f.write("\n")
        f.write(
            f'r{j+4},0,"^passport_expiry\[00000000{passport_no}\]$",{expiry},"",1,c119'
        )
        f.write("\n")
        f.write(f'r{j+5},0,"^birthday\[00000000{passport_no}\]$",{birthday},"",1,c119')
        f.write("\n")
        j = j + 6
